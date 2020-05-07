#!Python3

#import url requests and bs4
import bs4, requests
import os

dividendoutput = "divdata.xlsx"
dividendinput = "divsource.xlsx"



#Yahoo data function
def yahoodiv(produrl): 
    res = requests.get(produrl) #define the res variable that will go lookup the URL passed to it
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

    noresult = soup.select('span.D\(b\) > span:nth-child(1)')
    
    if noresult != []:
        print(noresult) #will error out if stock symbol is invalid
    
    #return each CSS selector as a list, stripping the stock name and symbol 
    revgrowth = soup.select('div.Pos\(r\):nth-child(4) > div:nth-child(1) > div:nth-child(1) > table:nth-child(2) > tbody:nth-child(1) > tr:nth-child(3) > td:nth-child(2)')
    payoutr = soup.select('div.Pstart\(20px\) > div:nth-child(3) > div:nth-child(1) > div:nth-child(1) > table:nth-child(2) > tbody:nth-child(1) > tr:nth-child(6) > td:nth-child(2)')
    oneyrret = soup.select('div.Pstart\(20px\) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > table:nth-child(2) > tbody:nth-child(1) > tr:nth-child(2) > td:nth-child(2)')
    divyield = soup.select('div.Pstart\(20px\) > div:nth-child(3) > div:nth-child(1) > div:nth-child(1) > table:nth-child(2) > tbody:nth-child(1) > tr:nth-child(2) > td:nth-child(2)')
    fiavg = soup.select('div.Pstart\(20px\) > div:nth-child(3) > div:nth-child(1) > div:nth-child(1) > table:nth-child(2) > tbody:nth-child(1) > tr:nth-child(5) > td:nth-child(2)')
    exdivdate = soup.select('div.Pstart\(20px\) > div:nth-child(3) > div:nth-child(1) > div:nth-child(1) > table:nth-child(2) > tbody:nth-child(1) > tr:nth-child(8) > td:nth-child(2)')
    paydate = soup.select('div.Pstart\(20px\) > div:nth-child(3) > div:nth-child(1) > div:nth-child(1) > table:nth-child(2) > tbody:nth-child(1) > tr:nth-child(7) > td:nth-child(2)')
    
        
    return [revgrowth[0].text.strip(), oneyrret[0].text.strip(), payoutr[0].text.strip(), divyield[0].text.strip(), fiavg[0].text.strip(), exdivdate[0].text.strip(), paydate[0].text.strip()]


#import symbols from funds spreadsheet
from openpyxl import load_workbook
from openpyxl.styles import Alignment
workbook = load_workbook(filename=dividendinput)
sheet = workbook.active

#set headings
sheet['A1'] = 'Name'
sheet['B1'] = 'Symbol'
sheet['C1'] = 'Yield'
sheet['D1'] = '1yr Revenue \nGrowth %'
sheet['E1'] = 'Payout \nRatio'
sheet['F1'] = '5yr Yield \nAvg'
sheet['G1'] = '1 Year \nReturn %'
headings = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']
for h in headings:
    sheet[h].alignment = Alignment(wrap_text=True)

#convert tuple result to a usable list, get symbol column only, starting at row 2 to avoid titles
stocklist = [item[0] for item in sheet.iter_rows(min_col=2,
                                                 max_col=2, min_row=2,
                                                 values_only=True)]
#define r for cell counter increase
r = 1

#for each stock symbol in the list create a custom URL for Yahoo function
#set variables for each item to return

for stock in stocklist:
    modiv = yahoodiv('https://finance.yahoo.com/quote/' + stock + '/key-statistics?p=' + stock)
    div_yld = modiv[3].split('%')[0]
    fiveyr_yld_avg = modiv[4].split('%')[0]
    payout = modiv[2].split('.')[0] + '%'
    payoutrc = modiv[2].split('%')[0]
    yr_ret = modiv[1].split('%')[0]
    revgrow = modiv[0].split('%')[0]
    r = r+1
    if r == 3:
        print('Getting dividend data')
    
    sheet['C' + str(r)] = div_yld
    sheet['D' + str(r)] = revgrow
    sheet['E' + str(r)] = payout
    sheet['F' + str(r)] = fiveyr_yld_avg
    sheet['G' + str(r)] = yr_ret



workbook.save(filename=dividendoutput)
print('Done with dividend data update')






